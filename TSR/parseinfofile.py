import collections
import re
import pandas as pd
import urllib
import os
import sys
import argparse
from datetime import date
from openpyxl import load_workbook

def saveData(df, variant, id):
    book = load_workbook('Data_TSRValidatedSignListMsg.xlsx')
    writer = pd.ExcelWriter('Data_TSRValidatedSignListMsg.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    df.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False, header= False)
    writer.save()
    writer.close()

def setName(dictionary, id, string):
    id = "id_{}".format(id)
    if id in dictionary:
        dictionary[id].append(string)
    else:
        dictionary.setdefault(id, [])
        dictionary[id].append(string)


def start(variant):
    
    # TSR variable definition
    TSRSignCountFlag = False
    TSRsignIDFlag = False
    TSRDictionaryList = {}
    Tsr_ValidatedSignListMsg = "40_Tsr_ValidatedSignListMsg"
    TSRHeaderFlag = True

    TSR_StringHolder = ['Frame index','Sign id',
                        'Mainsign id','MainSign type',
                        'Sign valid']

    with open(variant, 'r') as f:
        for line in f.readlines():
            nLine = line.strip()

            # For TSR ----
            if bool(re.match(r'SDD[\s]40_Tsr_ValidatedSignListMsg[\s]msgInfo[\s]frameindex[\s][0-9]+', nLine)):
                TSR_StringHolder[0] = nLine[40:]
            if bool(re.match(r'SDD[\s]40_Tsr_ValidatedSignListMsg[\s]data[\s]signCount[\s][^0]', nLine)):
                TSRSignCountFlag = True
            elif bool(re.match(r'SDD[\s]40_Tsr_ValidatedSignListMsg[\s]data[\s]signCount[\s][0]', nLine)):
                TSRSignCountFlag = False

            if TSRSignCountFlag:
                if bool(re.match(r'SDD[\s]40_Tsr_ValidatedSignListMsg[\s]data[\s]signs[\s][0-9]+[\s]id[\s][^0]+', nLine)):
                    TSR_StringHolder[1] = nLine[37:]
                    TSRsignIDFlag = True
                elif bool(re.match(r'SDD[\s]40_Tsr_ValidatedSignListMsg[\s]data[\s]signs[\s][0-9]+[\s]id[\s][0]+', nLine)):
                    TSRsignIDFlag = False

            if TSRsignIDFlag:
                if bool(re.match(r'SDD[\s]40_Tsr_ValidatedSignListMsg[\s]data[\s]signs[\s][0-9]+[\s]mainSign[\s]cat[\s][^0]+', nLine)):
                    TSR_StringHolder[2] = nLine[37:]
                elif bool(re.match(r'SDD[\s]40_Tsr_ValidatedSignListMsg[\s]data[\s]signs[\s][0-9]+[\s]mainSign[\s]type[\s][^0]+', nLine)):
                    TSR_StringHolder[3] = nLine[37:]
                else:
                    pass
    
                if bool(re.match(r'SDD[\s]40_Tsr_ValidatedSignListMsg[\s]data[\s]signs[\s][0-9]+[\s]valid[\s][1]', nLine)):
                    if TSRHeaderFlag:
                        TSRHeaderFlag = False
                        # Header
                        for ind in range(len(TSR_StringHolder)):
                            if ind == 0:
                                setName(TSRDictionaryList, ind, variant[31:49])
                            else:
                                setName(TSRDictionaryList, ind, " ")
                    else:
                        pass
                    TSR_StringHolder[4] = nLine[37:]
                    for ind in range(len(TSR_StringHolder)):
                        setName(TSRDictionaryList, ind, TSR_StringHolder[ind])
                else:
                    pass

    # TSR DataFrame Tsr_ValidatedSignListMsg
    if len(TSRDictionaryList) == 0:
        print("No valid data for " +  variant[:-4])
    else:
        TSRdf = pd.DataFrame({ key:pd.Series(value) for key, value in TSRDictionaryList.items() })
        TSRdf = TSRdf.drop_duplicates(subset='id_1', keep="first")
        saveData(TSRdf, variant, Tsr_ValidatedSignListMsg)
        print( Tsr_ValidatedSignListMsg + " Completed")
        
def main(argv):
    # Read arguments
    parser = argparse.ArgumentParser()
    parser.add_argument('-v', '--variant', dest="variant",
                        help='Variant: Input [Filename].txt')
    args = parser.parse_args()

    if args.variant:
        variant = args.variant
        print("Start")
        print("Calculating...")
        start(variant)
        print("End")
    else:
        print('Please input: python parseinfofile.py -v [filename].txt')
    
if __name__ == "__main__":
    main(sys.argv[1:])
